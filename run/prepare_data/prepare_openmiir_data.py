import sys
from omegaconf import DictConfig
import hydra
import einops
import json
from tqdm import tqdm
from openpyxl import load_workbook
import librosa
from mne import pick_types
from mne.preprocessing import ICA, read_ica
import numpy as np
from src.data_module.utils import check_ieeg, get_split, get_n_fold_split
import torch
import mne
mne.set_log_level('WARNING')
from pathlib import Path
import os
from src.utils.log import setup_logging, cprint, tracking
from src.data_module.utils import split_into_patches
import warnings

warnings.filterwarnings("ignore", category=FutureWarning, module="torch")


CONDITIONS = ['cued', 'non-cued', 'free']
STIMULUS_IDS = [1, 2, 3, 4, 11, 12, 13, 14, 21, 22, 23, 24]
STIMULUS_IDS2LABEL_ID = {
    1: 0, 2: 1, 3: 2, 4: 3, 11: 4, 12: 5, 13: 6, 14: 7, 21: 8, 22: 9, 23: 10, 24: 11
}
DEFAULT_VERSION = 1
sfreq = 512  # no down-sampling
subjects = ['P01','P04','P06','P07','P09','P11','P12','P13','P14']
LABEL_ID2SUBJECTS_IDS = {
    1: 'P01',
    2: 'P04',
    3: 'P06',
    4: 'P07',
    5: 'P09',
    6: 'P11',
    7: 'P12',
    8: 'P13',
    9: 'P14'
}
MASTOID_CHANNELS = [u'EXG5', u'EXG6']
stimuli = STIMULUS_IDS

include_cue = False
use_mastoid_reference = False

data_root = '/data/share/data/OpenMIIR/'
mne_data_root = os.path.join(data_root, 'eeg', 'mne')
ica_data_root = os.path.join(data_root, 'eeg', 'preprocessing', 'ica')

cache = dict()


def get_stimuli_version(subject):
    if subject in ['Pilot3','P01','P04','P05','P06','P07']:
        return 1
    else:
        return 2

def get_event_id(stimulus_id, condition):
    return stimulus_id * 10 + condition

def load_stimuli_metadata(data_root, version=None):

    if version is None:
        version = DEFAULT_VERSION

    xlsx_filepath = os.path.join(data_root, 'meta', 'Stimuli_Meta.v{}.xlsx'.format(version))
    workbook = load_workbook(filename=xlsx_filepath, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    meta = dict()
    for i in range(2, 14):
        stimulus_id = int(sheet.cell(i,1).value)
        meta[stimulus_id] = {
            'id' : stimulus_id,
            'label' : sheet.cell(i,2).value,
            'audio_file' : sheet.cell(i,3).value,
            'cue_file' : sheet.cell(i,3).value.replace('.wav', '_cue.wav'),
            'length_with_cue' : sheet.cell(i,4).value,
            'length_of_cue' : sheet.cell(i,5).value,
            'length_without_cue' : sheet.cell(i,6).value,
            'length_of_cue_only' : sheet.cell(i,7).value,
            'cue_bpm' : int(sheet.cell(i,8).value),
            'beats_per_bar' : int(sheet.cell(i,9).value),
            'num_bars' : int(sheet.cell(i,15).value),
            'cue_bars' : int(sheet.cell(i,16).value),
            'bpm' : int(sheet.cell(i,17).value),
            'approx_bar_length' : sheet.cell(i,12).value,
        }

        if version == 2:
            meta[stimulus_id]['bpm'] = meta[stimulus_id]['cue_bpm'] # use cue bpm

    return meta

def load_raw_info(subject,
             mne_data_root,
             verbose=False):
    mne_data_filepath = os.path.join(mne_data_root, '{}-raw.fif'.format(subject))
    raw = mne.io.Raw(mne_data_filepath, preload=False, verbose=verbose)
    return raw.info

def recording_has_mastoid_channels(subject):
    if subject in ['Pilot3','P01','P02','P03','P04','P05','P06','P07','P08']:
        return False
    else:
        return True

def merge_trial_and_audio_onsets(raw, use_audio_onsets=True, inplace=True, stim_channel='STI 014'):
    events = mne.find_events(raw, stim_channel='STI 014', shortest_event=0)
    merged = list()
    last_trial_event = None
    for i, event in enumerate(events):
        etype = event[2]
        if etype < 1000 or etype == 1111: # trial or noise onset
            if use_audio_onsets and events[i+1][2] == 1000: # followed by audio onset
                onset = events[i+1][0]
                merged.append([onset, 0, etype])
            else:
                # either we are not interested in audio onsets or there is none
                merged.append(event)
        # audio onsets (etype == 1000) are not copied
        if etype > 1111: # other events (keystrokes)
            merged.append(event)

    merged = np.asarray(merged, dtype=int)

    if inplace:
        stim_id = raw.ch_names.index(stim_channel)
        raw._data[stim_id,:].fill(0)     # delete data in stim channel
        raw.add_events(merged)

    return merged

def load_raw(subject, **args):
    return _load_raw(subject=subject, has_mastoid_channels=recording_has_mastoid_channels, **args)

def _load_raw(subject,
             mne_data_root=None,
             verbose=False,
             onsets=None,
             interpolate_bad_channels=False,
             has_mastoid_channels=None, # None=True, False, or callable(subject) returning True/False
             apply_reference=True, # by default, reference the data
             reference_mastoids=True):

    mne_data_filepath = os.path.join(mne_data_root, '{}-raw.fif'.format(subject))

    raw = mne.io.Raw(mne_data_filepath, preload=True, verbose=verbose)

    if apply_reference:
        if has_mastoid_channels is None \
            or has_mastoid_channels is True \
            or has_mastoid_channels(subject) is True:
            ## referencing to mastoids
            if reference_mastoids:
                mne.set_eeg_reference(raw, MASTOID_CHANNELS, copy=False) # inplace
            raw.drop_channels(MASTOID_CHANNELS)
        else:
            mne.set_eeg_reference(raw, copy=False)

    ## optional event merging
    if onsets == 'audio':
        merge_trial_and_audio_onsets(raw,
                                     use_audio_onsets=True,
                                     inplace=True,
                                     stim_channel='STI 014',
                                     )
    elif onsets == 'trials':
        merge_trial_and_audio_onsets(raw,
                                     use_audio_onsets=True,
                                     inplace=True,
                                     stim_channel='STI 014',
                                     )
    # else: keep both

    bads = raw.info['bads']
    if bads is not None and len(bads) > 0:
        if interpolate_bad_channels:
            raw.interpolate_bads()
    return raw

def fast_resample_mne(raw, sfreq, stim_picks=None, preserve_events=True, res_type='sinc_best', verbose=None):
    """Resample data channels.

    Resamples all channels. The data of the Raw object is modified inplace.

    The Raw object has to be constructed using preload=True (or string).

    WARNING: The intended purpose of this function is primarily to speed
    up computations (e.g., projection calculation) when precise timing
    of events is not required, as downsampling raw data effectively
    jitters trigger timings. It is generally recommended not to epoch
    downsampled data, but instead epoch and then downsample, as epoching
    downsampled data jitters triggers.

    Parameters
    ----------
    raw : nme raw object
        Raw data to filter.
    sfreq : float
        New sample rate to use.
    stim_picks : array of int | None
        Stim channels. These channels are simply subsampled or
        supersampled (without applying any filtering). This reduces
        resampling artifacts in stim channels, but may lead to missing
        triggers. If None, stim channels are automatically chosen using
        mne.pick_types(raw.info, meg=False, stim=True, exclude=[]).
    res_type : str
        If `scikits.samplerate` is installed, :func:`librosa.core.resample`
        will use ``res_type``. (Chooae between 'sinc_fastest', 'sinc_medium'
        and 'sinc_best' for the desired speed-vs-quality trade-off.)
        Otherwise, it will fall back on `scipy.signal.resample`
    verbose : bool, str, int, or None
        If not None, override default verbose level (see mne.verbose).
        Defaults to self.verbose.

    Notes
    -----
    For some data, it may be more accurate to use npad=0 to reduce
    artifacts. This is dataset dependent -- check your data!
    """
    self = raw  # this keeps the mne code intact

    if not self.preload:
        raise RuntimeError('Can only resample preloaded data')
    sfreq = float(sfreq)
    o_sfreq = float(self.info['sfreq'])

    offsets = np.concatenate(([0], np.cumsum(self._raw_lengths)))
    new_data = list()
    # set up stim channel processing
    if stim_picks is None:
        stim_picks = pick_types(self.info, meg=False, ref_meg=False,
                                stim=True, exclude=[])
    stim_picks = np.asanyarray(stim_picks)

    ### begin new code: save events in each stim channel ###
    if preserve_events:
        stim_events = dict()
        for sp in stim_picks:
            stim_channel_name = raw.ch_names[sp]
            stim_events[sp] = mne.find_events(raw, stim_channel=stim_channel_name,
                                              shortest_event=0, verbose=verbose)
    ### end new code: save events in each stim channel ###

    ratio = sfreq / o_sfreq
    for ri in range(len(self._raw_lengths)):
        data_chunk = self._data[:, offsets[ri]:offsets[ri + 1]]
        new_data_chunk = list()
        for i, channel in enumerate(data_chunk):
            new_data_chunk.append(librosa.resample(channel, orig_sr=o_sfreq, target_sr=sfreq, res_type=res_type))
        new_data_chunk = np.vstack(new_data_chunk)
        new_data.append(new_data_chunk)
        ### end changed code ###

        new_ntimes = new_data[ri].shape[1]

        # Now deal with the stim channels. In empirical testing, it was
        # faster to resample all channels (above) and then replace the
        # stim channels than it was to only resample the proper subset
        # of channels and then use np.insert() to restore the stims

        # figure out which points in old data to subsample
        # protect against out-of-bounds, which can happen (having
        # one sample more than expected) due to padding
        stim_inds = np.minimum(np.floor(np.arange(new_ntimes)
                                        / ratio).astype(int),
                               data_chunk.shape[1] - 1)
        for sp in stim_picks:
            new_data[ri][sp] = data_chunk[[sp]][:, stim_inds]

        self._first_samps[ri] = int(self._first_samps[ri] * ratio)
        self._last_samps[ri] = self._first_samps[ri] + new_ntimes - 1
        self._raw_lengths[ri] = new_ntimes

    # adjust affected variables
    self._data = np.concatenate(new_data, axis=1)
    # self.info['sfreq'] = sfreq
    raw.resample(sfreq)
    # self._update_times()

    ### begin new code: restore save events in each stim channel ###
    if preserve_events:
        for sp in stim_picks:
            raw._data[sp,:].fill(0)     # delete data in stim channel
            # scale onset times
            for event in stim_events[sp]:
                onset = int(np.floor(event[0] * ratio))
                event_id = event[2]
                if raw._data[sp,onset] > 0:
                    raw._data[sp,onset+1] = event_id
                else:
                    raw._data[sp,onset] = event_id


def resample_mne_events(events, o_sfreq, sfreq, fix_collisions=True):
    ratio = sfreq / o_sfreq
    resampled_events = list()
    for event in events:
        onset = int(np.floor(event[0] * ratio))
        event_id = event[2]

        if fix_collisions and \
            len(resampled_events) > 0 and \
            resampled_events[-1][0] == onset:
            onset += 1

        resampled_events.append([onset, 0, event_id])

    return np.asarray(resampled_events)

def load_ica(subject, description, ica_data_root=None):
    ica_filepath = os.path.join(ica_data_root,
                                '{}-{}-ica.fif'.format(subject, description))
    return read_ica(ica_filepath)

def load_and_preprocess_raw(subject,
                            onsets='audio',
                            interpolate_bad_channels=False,
                            reference_mastoids=True,
                            l_freq=0.5,
                            h_freq=30,
                            sfreq=None,
                            ica_cleaning=True,
                            ica_name='100p_64c',
                            l_freq2=None,
                            h_freq2=None,
                            verbose=None,
                            n_jobs=4,
                            mne_data_root=None,
                            ica_data_root=None,
                            ):

    # load the imported fif data, use the specified onsets
    raw = load_raw(subject,
                   onsets=onsets,
                   interpolate_bad_channels=interpolate_bad_channels,
                   reference_mastoids=reference_mastoids,
                   verbose=verbose,
                   mne_data_root=mne_data_root,
                   )

    # apply bandpass filter, use 4 processes to speed things up
    eeg_picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=False, stim=False)
    raw.filter(l_freq=l_freq, h_freq=h_freq, picks=eeg_picks, filter_length='10s',
               l_trans_bandwidth=0.5, h_trans_bandwidth=0.5, method='fft',
               n_jobs=n_jobs, verbose=verbose)

    # extract events
    # this comprises 240 trials, 60 noise events (1111) and 60 feedback events (2000=No, 2001=Yes)
    trial_events = mne.find_events(raw, stim_channel='STI 014', shortest_event=0, verbose=verbose)

    # resample data and eventa
    if sfreq is not None:
        orig_sfreq = raw.info['sfreq']
        if sfreq != orig_sfreq:
            fast_resample_mne(raw, sfreq, res_type='sinc_fastest', preserve_events=True, verbose=False)
            trial_events = resample_mne_events(trial_events, orig_sfreq, sfreq)

    if ica_cleaning:
        # load ica
        ica = load_ica(subject, description=ica_name, ica_data_root=ica_data_root)
        raw = ica.apply(raw, exclude=ica.exclude)

    if l_freq2 is not None or h_freq2 is not None:
        raw.filter(l_freq=l_freq2, h_freq=h_freq2, picks=eeg_picks, filter_length='10s',
               l_trans_bandwidth=0.1, h_trans_bandwidth=0.5, method='fft',
               n_jobs=n_jobs, verbose=verbose)

    return raw, trial_events


def gather_data(data_args, subject, pad = False):
    if subject in cache: return cache[subject]
    info = load_raw_info(subject, mne_data_root=mne_data_root)  # need to keep info with bad channels for later
    raw, trial_events = load_and_preprocess_raw(subject,
                                                mne_data_root=mne_data_root,
                                                ica_data_root=ica_data_root,
                                                onsets='audio',
                                                interpolate_bad_channels=False,
                                                reference_mastoids=use_mastoid_reference,
                                                l_freq=0.5,
                                                h_freq=30,
                                                sfreq=sfreq,
                                                ica_cleaning=False,
                                                verbose=None,
                                                n_jobs=8)
    cache[subject] = raw, trial_events, info
    # load metadata
    meta_version = get_stimuli_version(subject)
    meta = load_stimuli_metadata(data_root, meta_version)
    raw, trial_events, info = cache[subject]

    eeg_list = []
    mask_list = []
    label_list = []

    if data_args.dataset.task == 'perception':
        conditions = [1]
    elif data_args.dataset.task == 'imagination':
        conditions = [2, 3, 4]
    else:
        raise ValueError('Invalid task')

    for stim_id in tqdm(stimuli, desc = 'Stimulus'):
        if include_cue:
            trial_len = meta[stim_id]['length_with_cue']
        else:
            trial_len = meta[stim_id]['length_without_cue']
        trial_eeg_list = []
        for cond in conditions:
            event_id = get_event_id(stim_id, cond)
            # select EEG channels
            eeg_picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=False, stim=False, exclude=[])
            epochs = mne.Epochs(raw, events=trial_events, event_id=event_id,
                                tmin=0, tmax=trial_len,
                                proj=False, picks=eeg_picks, preload=True, verbose=False, baseline=(0, 0))
            # TEMPORARY FIX for update from mne-python 0.9 to 0.10
            # interpolate bad channels again as ICA will not be applied on them
            epochs.info['bads'] = info['bads']  # set bad channel flags for interpolation

            for bad_ch in epochs.info['bads']:
                if bad_ch in epochs.ch_names:
                    ch_idx = epochs.ch_names.index(bad_ch)
                    epochs._data[:, ch_idx, :] = 0

            for i, trial in enumerate(epochs.get_data()):
                trial_eeg_list.append(np.asarray(trial, dtype=np.float32))

        trial_eeg = np.stack(trial_eeg_list, axis=0)
        eeg = split_into_patches(trial_eeg, data_args.dataset.patch_len, data_args.dataset.patch_stride)
        eeg = einops.rearrange(eeg, 'b c s p -> (b s) c p')
        if pad:
            pad_size = data_args.dataset.input_channels - eeg.shape[1]
            if pad_size > 0:
                padded_ieeg = np.pad(eeg, ((0, 0), (0, pad_size),(0, 0)), mode='constant', constant_values=0)
                mask = np.ones((eeg.shape[0], data_args.dataset.input_channels))
                mask[:, eeg.shape[1]:] = 0
            else:
                padded_ieeg = eeg[:, :data_args.dataset.input_channels]
                mask = np.ones((eeg.shape[0], data_args.dataset.input_channels))
        else:
            padded_ieeg = eeg
            mask = np.ones((eeg.shape[0], eeg.shape[1]))

        eeg_list.append(padded_ieeg)
        mask_list.append(mask)
        label_list.append(np.ones(eeg.shape[0]) * STIMULUS_IDS2LABEL_ID[stim_id])

    eeg = np.concatenate(eeg_list, axis=0)
    mask = np.concatenate(mask_list, axis=0)
    label = np.concatenate(label_list, axis=0)

    return eeg, mask, label


@hydra.main(config_path="../conf", config_name="prepare_data", version_base="1.2")
def prepare_data(cfg: DictConfig):
    LOGGER = setup_logging(level = 20)
    processed_dir: Path = Path(cfg.dir.processed_dir) / cfg.dataset.name / cfg.dataset.task
    processed_dir.mkdir(parents=True, exist_ok=True)
    # if processed_dir.exists():
    #     shutil.rmtree(processed_dir)
    #     LOGGER.info_high(f"Removed dir: {processed_dir}")
    pad = cfg.dataset.pad
    id = cfg.dataset.id
    with tracking("Load and gather data", LOGGER):

        ieeg, mask, label = gather_data(cfg,subjects[int(id)-1], pad)

        LOGGER.info_high(f"Loaded data with shape: {ieeg.shape}")

    with tracking("Get and save split", LOGGER):
        if cfg.split_method == 'simple':
            train_split, eval_split, test_split = get_split(cfg, ieeg, label=label)

            train_split_filename = f'{cfg.dataset.name}_{id}_train_split.npy'
            eval_split_filename = f'{cfg.dataset.name}_{id}_eval_split.npy'
            test_split_filename = f'{cfg.dataset.name}_{id}_test_split.npy'

            np.save(processed_dir / train_split_filename, train_split)
            np.save(processed_dir / eval_split_filename, eval_split)
            np.save(processed_dir / test_split_filename, test_split)

        elif cfg.split_method == 'n_fold':
            fold_splits, test_indices = get_n_fold_split(cfg, ieeg, label=label)
            for fold_idx, (train_split, eval_split) in enumerate(fold_splits):
                train_split_filename = f'{cfg.dataset.name}_{id}_fold{fold_idx}_train_split.npy'
                eval_split_filename = f'{cfg.dataset.name}_{id}_fold{fold_idx}_eval_split.npy'
                test_split_filename = f'{cfg.dataset.name}_{id}_fold{fold_idx}_test_split.npy'

                np.save(processed_dir / train_split_filename, train_split)
                np.save(processed_dir / eval_split_filename, eval_split)
                np.save(processed_dir / test_split_filename, test_indices)


    with tracking("Prepare and save data", LOGGER):
        dataset_list = []
        for i in tqdm(range(len(ieeg)),desc='Preparing data'):
            data_dict = {
                'ieeg_raw_data': torch.tensor(ieeg[i]),
                'ieeg_mask'    : torch.tensor(mask[i]),
                'labels'       : torch.tensor(label[i]),
            }
            dataset_list.append(data_dict)
        torch.save(dataset_list, processed_dir / f'{cfg.dataset.name}_{id}_all_data.pt')

if __name__ == '__main__':
    prepare_data()